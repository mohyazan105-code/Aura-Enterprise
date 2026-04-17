"""
ActionAura Intelligence Engine
Pure-Python analytics: regression, K-Means, anomaly detection,
CLV, churn, ROI, attrition, process mining, legacy DB analysis.
No external ML dependencies required.
"""

import json
import math
import csv
import io
import sqlite3
import random
from datetime import datetime, timedelta
from database.db_manager import get_conn, get_records


# ─────────────────────────────────────────────────────────────────
# MATH PRIMITIVES
# ─────────────────────────────────────────────────────────────────

def _mean(lst):
    return sum(lst) / len(lst) if lst else 0.0

def _std(lst):
    if len(lst) < 2:
        return 0.0
    m = _mean(lst)
    return math.sqrt(sum((x - m) ** 2 for x in lst) / (len(lst) - 1))

def _zscore(value, mean, std):
    return (value - mean) / std if std else 0.0

def _covariance(x, y):
    mx, my = _mean(x), _mean(y)
    return sum((xi - mx) * (yi - my) for xi, yi in zip(x, y)) / max(len(x) - 1, 1)

def _pearson(x, y):
    sx, sy = _std(x), _std(y)
    return _covariance(x, y) / (sx * sy) if sx and sy else 0.0

def _distance(a, b):
    return math.sqrt(sum((ai - bi) ** 2 for ai, bi in zip(a, b)))


# ─────────────────────────────────────────────────────────────────
# LINEAR REGRESSION FORECASTER
# ─────────────────────────────────────────────────────────────────

def linear_regression(x, y):
    """Return (slope, intercept) using ordinary least squares."""
    n = len(x)
    if n < 2:
        return 0.0, _mean(y)
    sx = sum(x)
    sy = sum(y)
    sxy = sum(xi * yi for xi, yi in zip(x, y))
    sxx = sum(xi ** 2 for xi in x)
    denom = n * sxx - sx * sx
    if denom == 0:
        return 0.0, _mean(y)
    slope = (n * sxy - sx * sy) / denom
    intercept = (sy - slope * sx) / n
    return slope, intercept

def forecast_series(values, periods_ahead=6):
    """
    Forecast `periods_ahead` future values using linear regression.
    Returns: {
        historical: [...],
        forecast: [...],
        trend: 'up'|'down'|'stable',
        slope: float,
        confidence_band: float
    }
    """
    if not values:
        return {'historical': [], 'forecast': [], 'trend': 'stable', 'slope': 0, 'confidence_band': 0}

    x = list(range(len(values)))
    slope, intercept = linear_regression(x, values)

    forecast = []
    for i in range(1, periods_ahead + 1):
        xi = len(values) - 1 + i
        forecast.append(round(intercept + slope * xi, 2))

    residuals = [values[i] - (intercept + slope * i) for i in x]
    rmse = math.sqrt(_mean([r ** 2 for r in residuals])) if residuals else 0
    conf = round(rmse * 1.96, 2)  # 95% confidence band

    trend = 'up' if slope > 0.5 else ('down' if slope < -0.5 else 'stable')

    return {
        'historical': [round(v, 2) for v in values],
        'forecast': forecast,
        'trend': trend,
        'slope': round(slope, 4),
        'confidence_band': conf,
        'r_squared': round(_pearson(x, values) ** 2, 4) if len(values) > 2 else 0
    }


def moving_average(values, window=3):
    """Simple moving average smoothing."""
    if len(values) < window:
        return values
    result = []
    for i in range(len(values)):
        if i < window - 1:
            result.append(values[i])
        else:
            result.append(round(_mean(values[i - window + 1:i + 1]), 2))
    return result


# ─────────────────────────────────────────────────────────────────
# TIME SERIES ANALYSIS
# ─────────────────────────────────────────────────────────────────

def time_series_analysis(domain, department, metric='revenue', periods_ahead=6):
    """
    Full time-series: extract from DB → smooth → detect trend/seasonality → forecast.
    """
    rows, _ = get_records(domain, 'metrics', limit=500)
    dept_rows = sorted(
        [r for r in rows if r.get('department') == department and r.get('name') == metric],
        key=lambda r: r.get('date') or ''
    )

    if len(dept_rows) >= 4:
        values = [r['value'] for r in dept_rows]
        labels = [r.get('period') or (r.get('date') or '')[:7] for r in dept_rows]
    else:
        # Synthesize from transactions if no metrics records
        txn_rows, _ = get_records(domain, 'transactions', limit=500)
        monthly = {}
        for t in txn_rows:
            if t.get('type') == 'credit' and t.get('date'):
                period = t['date'][:7]
                monthly[period] = monthly.get(period, 0) + (t.get('amount') or 0)
        if monthly:
            sorted_m = sorted(monthly.items())
            labels = [k for k, _ in sorted_m]
            values = [v for _, v in sorted_m]
        else:
            now = datetime.now()
            labels = [(now - timedelta(days=30 * i)).strftime('%Y-%m') for i in range(11, -1, -1)]
            base = random.uniform(80000, 200000)
            values = [round(base + random.gauss(0, base * 0.08) + i * base * 0.02, 2) for i in range(12)]

    smoothed = moving_average(values, window=3)
    result = forecast_series(values, periods_ahead=periods_ahead)
    result['labels'] = labels
    result['smoothed'] = smoothed
    result['metric'] = metric
    result['department'] = department

    # Seasonality detection: variance by position mod 4 (quarters)
    if len(values) >= 8:
        quarters = {}
        for i, v in enumerate(values):
            q = i % 4
            quarters.setdefault(q, []).append(v)
        q_avgs = {q: _mean(vs) for q, vs in quarters.items()}
        overall_avg = _mean(values)
        seasonal_indices = {q: round(avg / overall_avg, 3) if overall_avg else 1 for q, avg in q_avgs.items()}
        result['seasonal_indices'] = seasonal_indices
        result['has_seasonality'] = max(seasonal_indices.values()) - min(seasonal_indices.values()) > 0.15
    else:
        result['has_seasonality'] = False

    return result


# ─────────────────────────────────────────────────────────────────
# ANOMALY DETECTION  (Z-score based)
# ─────────────────────────────────────────────────────────────────

def detect_anomalies(domain, department, threshold=2.5):
    """
    Detect anomalies in transactions and metrics using Z-score method.
    Returns list of anomalous records with their z-scores.
    """
    anomalies = []

    # Transactions
    txn_rows, _ = get_records(domain, 'transactions', limit=1000)
    amounts = [abs(r.get('amount') or 0) for r in txn_rows if r.get('amount')]
    if amounts:
        m, s = _mean(amounts), _std(amounts)
        for r in txn_rows:
            amt = abs(r.get('amount') or 0)
            z = _zscore(amt, m, s)
            if abs(z) >= threshold:
                anomalies.append({
                    'source': 'transaction',
                    'id': r.get('id'),
                    'date': r.get('date') or r.get('created_at', '')[:10],
                    'description': r.get('description', 'Transaction'),
                    'value': round(amt, 2),
                    'z_score': round(z, 2),
                    'severity': 'high' if abs(z) >= 3.5 else 'medium',
                    'direction': 'spike' if z > 0 else 'drop',
                    'department': r.get('department', department)
                })

    # Expenses
    exp_rows, _ = get_records(domain, 'expenses', limit=500)
    exp_amounts = [r.get('amount') or 0 for r in exp_rows if r.get('amount')]
    if exp_amounts:
        m, s = _mean(exp_amounts), _std(exp_amounts)
        for r in exp_rows:
            amt = r.get('amount') or 0
            z = _zscore(amt, m, s)
            if abs(z) >= threshold:
                anomalies.append({
                    'source': 'expense',
                    'id': r.get('id'),
                    'date': r.get('date') or r.get('created_at', '')[:10],
                    'description': r.get('description') or r.get('category', 'Expense'),
                    'value': round(amt, 2),
                    'z_score': round(z, 2),
                    'severity': 'high' if abs(z) >= 3.5 else 'medium',
                    'direction': 'spike' if z > 0 else 'drop',
                    'department': r.get('department', department)
                })

    anomalies.sort(key=lambda a: abs(a['z_score']), reverse=True)
    return anomalies[:20]


# ─────────────────────────────────────────────────────────────────
# K-MEANS CLUSTERING  (pure Python)
# ─────────────────────────────────────────────────────────────────

def kmeans(points, k=3, max_iter=100):
    """
    Real K-Means clustering.
    points: list of dicts with 'x', 'y' numeric fields.
    Returns points with 'cluster' field assigned.
    """
    if not points or k < 1:
        return points

    coords = [[p['x'], p['y']] for p in points]
    # Init centroids by spread sampling
    centroids = [coords[i * (len(coords) // k)] for i in range(k)]

    for _ in range(max_iter):
        # Assign clusters
        clusters = [[] for _ in range(k)]
        for i, c in enumerate(coords):
            dists = [_distance(c, ctr) for ctr in centroids]
            cluster_idx = dists.index(min(dists))
            clusters[cluster_idx].append(i)

        # Recompute centroids
        new_centroids = []
        for ci, members in enumerate(clusters):
            if members:
                new_x = _mean([coords[m][0] for m in members])
                new_y = _mean([coords[m][1] for m in members])
                new_centroids.append([new_x, new_y])
            else:
                new_centroids.append(centroids[ci])

        if new_centroids == centroids:
            break
        centroids = new_centroids

    # Final assignment
    result = []
    for i, p in enumerate(points):
        dists = [_distance(coords[i], ctr) for ctr in centroids]
        result.append({**p, 'cluster': dists.index(min(dists))})

    return result, centroids


def cluster_customers(domain, k=3):
    """Cluster customers by value and activity."""
    rows, _ = get_records(domain, 'customers', limit=500)
    if not rows:
        return {'points': [], 'centroids': [], 'k': k, 'labels': ['Segment A', 'Segment B', 'Segment C']}

    # Normalize to 0-100
    values = [r.get('value') or 0 for r in rows]
    max_v = max(values) or 1

    points = []
    for r in rows:
        val = (r.get('value') or 0) / max_v * 100
        # Use id as proxy for recency (higher id = more recent)
        recency = (r.get('id') or 1) / max((r2.get('id') or 1) for r2 in rows) * 100
        points.append({
            'x': round(val, 2),
            'y': round(recency, 2),
            'label': r.get('name', 'Customer'),
            'id': r.get('id')
        })

    clustered, centroids = kmeans(points, k=min(k, len(points)))
    colors = ['#1a73e8', '#00e676', '#f57c00', '#e53935', '#8e24aa']
    segment_labels = {0: 'High Value', 1: 'Growing', 2: 'At Risk', 3: 'Dormant', 4: 'New'}

    return {
        'points': clustered,
        'centroids': [{'x': round(c[0], 2), 'y': round(c[1], 2)} for c in centroids],
        'k': k,
        'labels': [segment_labels.get(i, f'Segment {i+1}') for i in range(k)],
        'colors': colors[:k],
        'axes': {'x': 'Customer Value (normalized)', 'y': 'Recency Score'}
    }


def cluster_employees(domain, k=3):
    """Cluster employees by salary and performance."""
    rows, _ = get_records(domain, 'employees', limit=500)
    perf_rows, _ = get_records(domain, 'performance', limit=500)
    perf_map = {r.get('employee_id'): r.get('score') or 70 for r in perf_rows}

    if not rows:
        return {'points': [], 'centroids': [], 'k': k}

    salaries = [r.get('salary') or 0 for r in rows]
    max_s = max(salaries) or 1

    points = []
    for r in rows:
        sal_norm = (r.get('salary') or 0) / max_s * 100
        perf = perf_map.get(r.get('id'), random.uniform(50, 90))
        points.append({'x': round(sal_norm, 2), 'y': round(perf, 2), 'label': r.get('name', 'Employee')})

    clustered, centroids = kmeans(points, k=min(k, len(points)))
    return {
        'points': clustered,
        'centroids': [{'x': round(c[0], 2), 'y': round(c[1], 2)} for c in centroids],
        'k': k,
        'labels': ['High Performers', 'Core Team', 'Needs Support'][:k],
        'colors': ['#43a047', '#1a73e8', '#e53935'][:k],
        'axes': {'x': 'Salary Level (normalized)', 'y': 'Performance Score'}
    }


# ─────────────────────────────────────────────────────────────────
# CUSTOMER LIFETIME VALUE
# ─────────────────────────────────────────────────────────────────

def calculate_clv(domain):
    """
    CLV = Avg Purchase Value × Purchase Frequency × Customer Lifespan
    """
    customers, _ = get_records(domain, 'customers', limit=500)
    deals, _ = get_records(domain, 'deals', limit=500)
    invoices, _ = get_records(domain, 'invoices', limit=500)

    result = []
    for c in customers:
        cid = c.get('id')
        cust_deals = [d for d in deals if d.get('customer_id') == cid]
        cust_invoices = [i for i in invoices if i.get('client') == c.get('name')]

        total_value = c.get('value') or sum(d.get('value') or 0 for d in cust_deals)
        n_transactions = len(cust_deals) + len(cust_invoices) or 1

        # Estimate lifespan in years from created_at
        created = c.get('created_at') or ''
        try:
            days_active = (datetime.now() - datetime.fromisoformat(created[:10])).days
        except Exception:
            days_active = 365

        lifespan_years = max(days_active / 365, 0.1)
        avg_purchase = total_value / n_transactions
        frequency = n_transactions / lifespan_years
        clv = round(avg_purchase * frequency * 3, 2)  # 3-year projection

        result.append({
            'customer': c.get('name'),
            'id': cid,
            'status': c.get('status', 'active'),
            'clv': clv,
            'total_value': round(total_value, 2),
            'transactions': n_transactions,
            'lifespan_years': round(lifespan_years, 1),
            'tier': 'Platinum' if clv > 50000 else ('Gold' if clv > 20000 else ('Silver' if clv > 5000 else 'Bronze'))
        })

    result.sort(key=lambda c:  c['clv'], reverse=True)
    avg_clv = round(_mean([r['clv'] for r in result]), 2) if result else 0
    return {'customers': result[:50], 'avg_clv': avg_clv, 'total_customers': len(result)}


# ─────────────────────────────────────────────────────────────────
# CHURN PREDICTION
# ─────────────────────────────────────────────────────────────────

def churn_prediction(domain):
    """Rule-based churn risk scoring."""
    customers, _ = get_records(domain, 'customers', limit=500)
    cases, _ = get_records(domain, 'cases', limit=500)
    deals, _ = get_records(domain, 'deals', limit=500)

    result = []
    for c in customers:
        cid = c.get('id')
        score = 0  # 0-100 churn risk

        # Factor 1: inactive status
        if c.get('status') == 'inactive':
            score += 40

        # Factor 2: open/unresolved cases
        open_cases = [cs for cs in cases if cs.get('customer_id') == cid and cs.get('status') == 'open']
        score += min(len(open_cases) * 10, 30)

        # Factor 3: no recent deals
        cust_deals = [d for d in deals if d.get('customer_id') == cid]
        if not cust_deals:
            score += 20
        elif all(d.get('stage') == 'closed-lost' for d in cust_deals):
            score += 15

        # Factor 4: low value
        if (c.get('value') or 0) < 1000:
            score += 10

        score = min(score, 100)
        result.append({
            'customer': c.get('name'),
            'id': cid,
            'churn_risk': score,
            'risk_level': 'Critical' if score >= 60 else ('High' if score >= 40 else ('Medium' if score >= 20 else 'Low')),
            'status': c.get('status', 'active'),
            'open_cases': len(open_cases),
            'value': c.get('value') or 0
        })

    result.sort(key=lambda r: r['churn_risk'], reverse=True)
    high_risk = [r for r in result if r['churn_risk'] >= 40]
    return {'customers': result[:50], 'high_risk_count': len(high_risk), 'total': len(result)}


# ─────────────────────────────────────────────────────────────────
# ATTRITION RISK (HR)
# ─────────────────────────────────────────────────────────────────

def attrition_risk(domain):
    """Score each employee's flight risk."""
    employees, _ = get_records(domain, 'employees', limit=500)
    leaves, _ = get_records(domain, 'leaves', limit=500)
    perf, _ = get_records(domain, 'performance', limit=500)
    perf_map = {r.get('employee_id'): r.get('score') or 70 for r in perf}

    result = []
    for e in employees:
        eid = e.get('id')
        score = 0

        # Amount of leave taken
        emp_leaves = [l for l in leaves if l.get('employee_id') == eid]
        score += min(len(emp_leaves) * 5, 25)

        # Low performance
        p = perf_map.get(eid, 75)
        if p < 60:
            score += 30
        elif p < 75:
            score += 15

        # Tenure — very new or very long = higher risk
        try:
            hire_date = datetime.fromisoformat((e.get('hire_date') or '2020-01-01')[:10])
            tenure_days = (datetime.now() - hire_date).days
            if tenure_days < 180 or tenure_days > 2000:
                score += 15
        except Exception:
            pass

        # On leave now
        if e.get('status') == 'on-leave':
            score += 10

        score = min(score, 100)
        result.append({
            'employee': e.get('name'),
            'id': eid,
            'department': e.get('department'),
            'position': e.get('position'),
            'attrition_risk': score,
            'risk_level': 'Critical' if score >= 60 else ('High' if score >= 40 else ('Medium' if score >= 20 else 'Low')),
            'performance_score': perf_map.get(eid, 'N/A'),
            'leave_count': len(emp_leaves)
        })

    result.sort(key=lambda r: r['attrition_risk'], reverse=True)
    return {'employees': result[:50], 'high_risk_count': len([r for r in result if r['attrition_risk'] >= 40]), 'total': len(result)}


# ─────────────────────────────────────────────────────────────────
# CAMPAIGN ROI ANALYSIS
# ─────────────────────────────────────────────────────────────────

def campaign_roi_analysis(domain):
    """Real ROI calculation from campaigns table."""
    campaigns, _ = get_records(domain, 'campaigns', limit=200)
    result = []
    for c in campaigns:
        budget = c.get('budget') or 0
        spent = c.get('spent') or 0
        leads = c.get('leads_generated') or 0
        roi_stored = c.get('roi') or 0

        # Calculate ROI: (Revenue - Cost) / Cost × 100
        # Estimate revenue from leads (avg deal ~$5,000)
        est_revenue = leads * 5000 * 0.15  # 15% conversion assumption
        roi = round((est_revenue - spent) / spent * 100, 1) if spent > 0 else 0

        result.append({
            'name': c.get('name'),
            'type': c.get('type'),
            'budget': budget,
            'spent': spent,
            'leads': leads,
            'status': c.get('status'),
            'roi_pct': roi,
            'est_revenue': round(est_revenue, 2),
            'efficiency': round(leads / spent * 1000, 2) if spent else 0,  # leads per $1000
            'budget_utilization': round(spent / budget * 100, 1) if budget else 0
        })

    result.sort(key=lambda r: r['roi_pct'], reverse=True)
    avg_roi = round(_mean([r['roi_pct'] for r in result]), 1) if result else 0
    return {'campaigns': result, 'avg_roi': avg_roi, 'total_spend': round(sum(r['spent'] for r in result), 2)}


# ─────────────────────────────────────────────────────────────────
# BUDGET VS ACTUAL ANALYSIS
# ─────────────────────────────────────────────────────────────────

def budget_vs_actual(domain):
    """Real budget deviation using budgets + expenses tables."""
    budgets, _ = get_records(domain, 'budgets', limit=200)
    expenses, _ = get_records(domain, 'expenses', limit=500)

    by_dept = {}
    for b in budgets:
        dept = b.get('department', 'unknown')
        by_dept.setdefault(dept, {'allocated': 0, 'spent_db': 0, 'budget_count': 0})
        by_dept[dept]['allocated'] += b.get('allocated') or 0
        by_dept[dept]['spent_db'] += b.get('spent') or 0
        by_dept[dept]['budget_count'] += 1

    # Supplement with actual expenses
    for e in expenses:
        dept = e.get('department', 'unknown')
        if dept in by_dept and e.get('status') == 'approved':
            by_dept[dept]['actual_expenses'] = by_dept[dept].get('actual_expenses', 0) + (e.get('amount') or 0)

    result = []
    for dept, data in by_dept.items():
        alloc = data['allocated']
        spent = data.get('actual_expenses') or data['spent_db']
        remaining = alloc - spent
        utilization = round(spent / alloc * 100, 1) if alloc else 0
        variance = round(remaining / alloc * 100, 1) if alloc else 0
        result.append({
            'department': dept,
            'allocated': round(alloc, 2),
            'spent': round(spent, 2),
            'remaining': round(remaining, 2),
            'utilization_pct': utilization,
            'variance_pct': variance,
            'status': 'over-budget' if spent > alloc else ('at-risk' if utilization > 85 else 'on-track')
        })

    result.sort(key=lambda r: r['utilization_pct'], reverse=True)
    return {
        'departments': result,
        'total_allocated': round(sum(r['allocated'] for r in result), 2),
        'total_spent': round(sum(r['spent'] for r in result), 2),
        'over_budget_depts': [r['department'] for r in result if r['status'] == 'over-budget']
    }


# ─────────────────────────────────────────────────────────────────
# PROCESS MINING
# ─────────────────────────────────────────────────────────────────

def process_mining(domain, department):
    """Analyze workflow logs to discover real process patterns."""
    instances, _ = get_records(domain, 'workflow_instances', limit=500)
    definitions, _ = get_records(domain, 'workflow_definitions', limit=50)

    def_map = {d['id']: d for d in definitions if d.get('department') == department}
    dept_instances = [i for i in instances if i.get('definition_id') in def_map]

    if not dept_instances:
        dept_instances = instances[:20]  # fallback to all

    # Completion rates
    total = len(dept_instances)
    completed = len([i for i in dept_instances if i.get('status') == 'completed'])
    rejected = len([i for i in dept_instances if i.get('status') == 'rejected'])
    pending = len([i for i in dept_instances if i.get('status') in ('pending', 'active')])

    # Average steps reached
    avg_step = _mean([i.get('current_step') or 1 for i in dept_instances])

    # Bottleneck detection: find which step has most pending
    pending_by_step = {}
    for i in dept_instances:
        if i.get('status') in ('pending', 'active'):
            step = i.get('current_step') or 1
            pending_by_step[step] = pending_by_step.get(step, 0) + 1

    bottleneck_step = max(pending_by_step, key=pending_by_step.get) if pending_by_step else None

    # Process variants: unique paths
    paths = {}
    for i in dept_instances:
        try:
            hist = json.loads(i.get('history_json') or '[]')
            path_key = '→'.join([h.get('action', '?') for h in hist[:4]])
        except Exception:
            path_key = i.get('status', 'unknown')
        paths[path_key] = paths.get(path_key, 0) + 1

    variants = sorted([{'path': k, 'count': v, 'frequency_pct': round(v / total * 100, 1)} for k, v in paths.items()], key=lambda x: x['count'], reverse=True)

    return {
        'total_instances': total,
        'completed': completed,
        'rejected': rejected,
        'pending': pending,
        'completion_rate': round(completed / total * 100, 1) if total else 0,
        'rejection_rate': round(rejected / total * 100, 1) if total else 0,
        'avg_step_reached': round(avg_step, 1),
        'bottleneck_step': bottleneck_step,
        'pending_by_step': pending_by_step,
        'process_variants': variants[:10],
        'insights': _process_insights(completed, rejected, total, bottleneck_step)
    }

def _process_insights(completed, rejected, total, bottleneck_step):
    insights = []
    if total == 0:
        return ['No process data available yet.']
    comp_rate = completed / total * 100
    if comp_rate < 60:
        insights.append(f'⚠️ Completion rate is {comp_rate:.0f}% — below the 80% target threshold.')
    else:
        insights.append(f'✅ Healthy completion rate of {comp_rate:.0f}%.')
    if rejected / total > 0.2:
        insights.append(f'🚨 High rejection rate ({rejected/total*100:.0f}%) — review approval criteria.')
    if bottleneck_step:
        insights.append(f'🔴 Bottleneck detected at Step {bottleneck_step} — most processes stall here.')
    return insights


# ─────────────────────────────────────────────────────────────────
# ADAPTIVE KPI ENGINE  (replaces hardcoded get_dashboard_kpis)
# ─────────────────────────────────────────────────────────────────

def get_adaptive_kpis(domain, department, role):
    """
    Returns KPIs computed from live DB data.
    Falls back to intelligent defaults enriched with real counts.
    """
    conn = get_conn(domain)
    kpis = []

    try:
        cur = conn.cursor()

        if department == 'finance' or department == 'analytics':
            # Real revenue KPI
            cur.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='credit'")
            revenue = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='debit'")
            expenses = abs(cur.fetchone()[0])
            margin = round((revenue - expenses) / revenue * 100, 1) if revenue else 0

            cur.execute("SELECT COUNT(*) FROM invoices WHERE status='overdue'")
            overdue = cur.fetchone()[0]

            cur.execute("SELECT COALESCE(SUM(allocated),0), COALESCE(SUM(spent),0) FROM budgets")
            alloc, spent = cur.fetchone()
            budget_util = round(spent / alloc * 100, 1) if alloc else 0

            kpis = [
                {'id': 'fin_rev', 'label': 'Total Revenue', 'value': round(revenue, 0), 'unit': '$', 'icon': '💰', 'color': '#43a047', 'trend': f'{margin:+.1f}% margin', 'trend_dir': 'up' if margin > 0 else 'down', 'type': 'kpi'},
                {'id': 'fin_budget', 'label': 'Budget Utilization', 'value': budget_util, 'unit': '%', 'icon': '📊', 'color': '#1e88e5', 'trend': f'${round(alloc - spent):,.0f} remaining', 'trend_dir': 'neutral' if budget_util < 85 else 'down', 'type': 'kpi'},
                {'id': 'fin_overdue', 'label': 'Overdue Invoices', 'value': overdue, 'unit': 'items', 'icon': '⚠️', 'color': '#e53935' if overdue > 0 else '#43a047', 'trend': 'Needs follow-up' if overdue > 0 else 'All clear', 'trend_dir': 'down' if overdue > 0 else 'up', 'type': 'kpi'},
            ]

        elif department == 'hr':
            cur.execute("SELECT COUNT(*) FROM employees WHERE status='active'")
            active = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM employees WHERE status='on-leave'")
            on_leave = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM employees")
            total = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(AVG(score),0) FROM performance")
            avg_perf = round(cur.fetchone()[0], 1)
            retention = round(active / total * 100, 1) if total else 0

            kpis = [
                {'id': 'hr_active', 'label': 'Active Headcount', 'value': active, 'unit': 'staff', 'icon': '👥', 'color': '#1e88e5', 'trend': f'{total} total', 'trend_dir': 'neutral', 'type': 'kpi'},
                {'id': 'hr_ret', 'label': 'Retention Rate', 'value': retention, 'unit': '%', 'icon': '🛡️', 'color': '#43a047' if retention > 90 else '#f57c00', 'trend': f'{on_leave} on leave', 'trend_dir': 'up' if retention > 90 else 'down', 'type': 'kpi'},
                {'id': 'hr_perf', 'label': 'Avg Performance', 'value': avg_perf, 'unit': '/100', 'icon': '⭐', 'color': '#8e24aa', 'trend': 'Latest review cycle', 'trend_dir': 'up' if avg_perf >= 75 else 'down', 'type': 'kpi'},
            ]

        elif department == 'crm' or department == 'marketing':
            cur.execute("SELECT COUNT(*) FROM customers WHERE status='active'")
            active_cust = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM deals WHERE stage='closed-won'")
            won = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM deals")
            total_deals = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM leads WHERE status='qualified'")
            qualified = cur.fetchone()[0]
            conv = round(won / total_deals * 100, 1) if total_deals else 0
            cur.execute("SELECT COALESCE(AVG(roi),0) FROM campaigns")
            avg_roi = round(cur.fetchone()[0], 1)

            kpis = [
                {'id': 'crm_active', 'label': 'Active Customers', 'value': active_cust, 'unit': '', 'icon': '🤝', 'color': '#fb8c00', 'trend': f'{qualified} qualified leads', 'trend_dir': 'neutral', 'type': 'kpi'},
                {'id': 'crm_conv', 'label': 'Deal Conversion', 'value': conv, 'unit': '%', 'icon': '🎯', 'color': '#43a047' if conv > 20 else '#e53935', 'trend': f'{won}/{total_deals} deals won', 'trend_dir': 'up' if conv > 20 else 'down', 'type': 'kpi'},
                {'id': 'crm_roi', 'label': 'Campaign ROI', 'value': avg_roi, 'unit': 'x', 'icon': '📣', 'color': '#1e88e5', 'trend': 'Marketing efficiency', 'trend_dir': 'up' if avg_roi > 2 else 'neutral', 'type': 'kpi'},
            ]

        elif department == 'pm' or department == 'operations':
            cur.execute("SELECT COUNT(*) FROM projects WHERE status='active'")
            active_proj = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(AVG(completion),0) FROM projects WHERE status='active'")
            avg_comp = round(cur.fetchone()[0], 1)
            cur.execute("SELECT COUNT(*) FROM tasks WHERE status='done'")
            done_tasks = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM tasks")
            total_tasks = cur.fetchone()[0]
            task_rate = round(done_tasks / total_tasks * 100, 1) if total_tasks else 0

            kpis = [
                {'id': 'pm_proj', 'label': 'Active Projects', 'value': active_proj, 'unit': '', 'icon': '📋', 'color': '#00897b', 'trend': f'Avg {avg_comp}% complete', 'trend_dir': 'neutral', 'type': 'kpi'},
                {'id': 'pm_task', 'label': 'Task Completion Rate', 'value': task_rate, 'unit': '%', 'icon': '✅', 'color': '#43a047' if task_rate > 70 else '#f57c00', 'trend': f'{done_tasks}/{total_tasks} tasks', 'trend_dir': 'up' if task_rate > 70 else 'down', 'type': 'kpi'},
                {'id': 'pm_eff', 'label': 'Project Completion', 'value': avg_comp, 'unit': '%', 'icon': '⚡', 'color': '#1e88e5', 'trend': 'Avg across active', 'trend_dir': 'up', 'type': 'kpi'},
            ]

        elif department == 'logistics':
            cur.execute("SELECT COUNT(*) FROM inventory WHERE quantity <= min_stock")
            low_stock = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM shipments WHERE status='delivered'")
            delivered = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM shipments")
            total_ship = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(AVG(reliability_score),0) FROM suppliers")
            supplier_rel = round(cur.fetchone()[0], 1)
            on_time = round(delivered / total_ship * 100, 1) if total_ship else 0

            kpis = [
                {'id': 'log_stock', 'label': 'Low Stock Alerts', 'value': low_stock, 'unit': 'items', 'icon': '📦', 'color': '#e53935' if low_stock > 0 else '#43a047', 'trend': 'Below minimum threshold', 'trend_dir': 'down' if low_stock > 0 else 'neutral', 'type': 'kpi'},
                {'id': 'log_del', 'label': 'On-Time Delivery', 'value': on_time, 'unit': '%', 'icon': '🚚', 'color': '#43a047' if on_time > 90 else '#f57c00', 'trend': f'{delivered}/{total_ship} shipments', 'trend_dir': 'up' if on_time > 90 else 'down', 'type': 'kpi'},
                {'id': 'log_sup', 'label': 'Supplier Reliability', 'value': supplier_rel, 'unit': '', 'icon': '🤝', 'color': '#1e88e5', 'trend': 'Reliability index', 'trend_dir': 'neutral', 'type': 'kpi'},
            ]

        else:
            # Generic fallback from actual DB counts
            cur.execute("SELECT COUNT(*) FROM employees WHERE status='active'")
            emp_count = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM tasks WHERE status='done'")
            done_t = cur.fetchone()[0]
            cur.execute("SELECT COUNT(*) FROM tasks")
            tot_t = cur.fetchone()[0]
            cur.execute("SELECT COALESCE(SUM(amount),0) FROM transactions WHERE type='credit'")
            rev = round(cur.fetchone()[0], 0)
            kpis = [
                {'id': 'gen_emp', 'label': 'Active Staff', 'value': emp_count, 'unit': '', 'icon': '👥', 'color': '#1e88e5', 'trend': 'All domains', 'trend_dir': 'neutral', 'type': 'kpi'},
                {'id': 'gen_task', 'label': 'Task Completion', 'value': round(done_t / tot_t * 100, 1) if tot_t else 0, 'unit': '%', 'icon': '✅', 'color': '#43a047', 'trend': f'{done_t}/{tot_t} tasks', 'trend_dir': 'up', 'type': 'kpi'},
                {'id': 'gen_rev', 'label': 'Total Revenue', 'value': rev, 'unit': '$', 'icon': '💰', 'color': '#8e24aa', 'trend': 'All transactions', 'trend_dir': 'neutral', 'type': 'kpi'},
            ]

        # Append custom KPIs for this user
        cur.execute("SELECT * FROM kpi_custom WHERE domain=? AND department=? AND is_active=1 AND (role='all' OR role=?)", (domain, department, role))
        custom = [dict(zip([d[0] for d in cur.description], row)) for row in cur.fetchall()]
        for ck in custom:
            kpis.append({
                'id': f'custom_{ck["id"]}',
                'label': ck['name'],
                'value': ck.get('target', 0),
                'unit': ck.get('unit', '%'),
                'icon': ck.get('icon', '📊'),
                'color': ck.get('color', '#1a73e8'),
                'trend': ck.get('formula', 'Custom KPI'),
                'trend_dir': 'neutral',
                'type': 'custom',
                'is_custom': True
            })

    except Exception as e:
        # Safe fallback
        kpis = [{'id': 'err', 'label': 'KPI Engine Loading', 'value': '...', 'unit': '', 'icon': '⚙️', 'color': '#1e88e5', 'trend': str(e)[:40], 'trend_dir': 'neutral', 'type': 'kpi'}]
    finally:
        conn.close()

    return kpis


# ─────────────────────────────────────────────────────────────────
# LEARNING LOOP  — saves patterns to ai_patterns table
# ─────────────────────────────────────────────────────────────────

def learn_from_data(domain):
    """
    Master learning loop. Runs after every data write.
    Scans transactions, employees, customers, tasks → persists patterns.
    """
    from database.db_manager import DEPARTMENTS
    conn = get_conn(domain)
    cur = conn.cursor()

    try:
        now = datetime.now().isoformat()

        for dept in DEPARTMENTS.keys():
            # 1. Revenue trend pattern
            try:
                ts = time_series_analysis(domain, dept, 'revenue', periods_ahead=6)
                _upsert_pattern(cur, domain, dept, 'trend', 'revenue', ts, ts.get('r_squared', 0))
            except Exception:
                pass

            # 2. Anomaly summary
            try:
                anomalies = detect_anomalies(domain, dept)
                _upsert_pattern(cur, domain, dept, 'anomaly', 'transactions',
                                {'count': len(anomalies), 'top': anomalies[:3]},
                                0.95)
            except Exception:
                pass

        # 3. Customer cluster pattern
        try:
            clusters = cluster_customers(domain)
            _upsert_pattern(cur, domain, None, 'cluster', 'customers', clusters, 0.90)
        except Exception:
            pass

        # 4. Save forecast predictions
        try:
            ts = time_series_analysis(domain, 'finance', 'revenue', periods_ahead=6)
            now_dt = datetime.now()
            cur.execute("DELETE FROM ai_predictions WHERE domain=? AND metric='revenue'", (domain,))
            for i, val in enumerate(ts.get('forecast', [])):
                period = (now_dt + timedelta(days=30 * (i + 1))).strftime('%Y-%m')
                band = ts.get('confidence_band', 0)
                cur.execute("""
                    INSERT INTO ai_predictions (domain, department, metric, period, predicted_value, confidence_low, confidence_high, model)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (domain, 'finance', 'revenue', period, val, max(0, val - band), val + band, 'linear'))
        except Exception:
            pass

        conn.commit()
    finally:
        conn.close()


def _upsert_pattern(cur, domain, department, pattern_type, metric, value, confidence):
    """Insert or update a pattern record."""
    now = datetime.now().isoformat()
    cur.execute("""
        SELECT id FROM ai_patterns WHERE domain=? AND department IS ? AND pattern_type=? AND metric=?
    """, (domain, department, pattern_type, metric))
    existing = cur.fetchone()
    val_json = json.dumps(value, default=str)
    if existing:
        cur.execute("""
            UPDATE ai_patterns SET value_json=?, confidence=?, updated_at=? WHERE id=?
        """, (val_json, confidence, now, existing[0]))
    else:
        cur.execute("""
            INSERT INTO ai_patterns (domain, department, pattern_type, metric, value_json, confidence, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (domain, department, pattern_type, metric, val_json, confidence, now, now))


def get_patterns(domain):
    """Retrieve all learned patterns for a domain."""
    conn = get_conn(domain)
    cur = conn.cursor()
    try:
        cur.execute("SELECT * FROM ai_patterns WHERE domain=? ORDER BY updated_at DESC", (domain,))
        cols = [d[0] for d in cur.description]
        rows = [dict(zip(cols, r)) for r in cur.fetchall()]
        for r in rows:
            try:
                r['value_json'] = json.loads(r['value_json'])
            except Exception:
                pass
        return rows
    finally:
        conn.close()


# ─────────────────────────────────────────────────────────────────
# LEGACY DATABASE ANALYZER  (SQLite + CSV + Excel)
# ─────────────────────────────────────────────────────────────────

ENTITY_FINGERPRINTS = {
    'customer':  ['customer', 'client', 'buyer', 'consumer', 'user', 'account'],
    'employee':  ['employee', 'staff', 'worker', 'personnel', 'hr', 'payroll'],
    'product':   ['product', 'item', 'sku', 'goods', 'inventory', 'stock'],
    'order':     ['order', 'sale', 'purchase', 'transaction', 'invoice'],
    'supplier':  ['supplier', 'vendor', 'provider', 'partner'],
    'project':   ['project', 'task', 'milestone', 'sprint', 'ticket'],
    'financial': ['budget', 'expense', 'revenue', 'cost', 'payment', 'amount', 'price'],
}

DEPT_SUGGESTIONS = {
    'customer': 'CRM',  'employee': 'HR', 'product': 'Logistics',
    'order': 'Finance', 'supplier': 'Logistics', 'project': 'Project Management',
    'financial': 'Finance',
}


def analyze_legacy_db(file_bytes, filename):
    """
    Analyze an uploaded legacy database file.
    Supports: .db (SQLite), .csv, .xlsx
    Returns full analysis report.
    """
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

    if ext == 'db':
        return _analyze_sqlite(file_bytes)
    elif ext == 'csv':
        return _analyze_csv(file_bytes)
    elif ext in ('xlsx', 'xls'):
        return _analyze_excel(file_bytes)
    else:
        return {'error': f'Unsupported file type: .{ext}. Supported: .db, .csv, .xlsx'}


def _profile_columns(columns, sample_rows):
    """Profile each column: type, null %, unique count, range."""
    profiles = []
    for col in columns:
        values = [r.get(col) if isinstance(r, dict) else r[columns.index(col)] for r in sample_rows]
        non_null = [v for v in values if v is not None and v != '']
        numerics = []
        for v in non_null:
            try:
                numerics.append(float(v))
            except Exception:
                pass
        profiles.append({
            'column': col,
            'null_pct': round((1 - len(non_null) / len(values)) * 100, 1) if values else 0,
            'unique_count': len(set(str(v) for v in non_null)),
            'is_numeric': len(numerics) > len(non_null) * 0.7,
            'min': round(min(numerics), 2) if numerics else None,
            'max': round(max(numerics), 2) if numerics else None,
            'mean': round(_mean(numerics), 2) if numerics else None,
            'inferred_type': 'numeric' if numerics else ('date' if any(k in col.lower() for k in ['date', 'at', 'time']) else 'text'),
            'is_likely_fk': col.lower().endswith('_id') or col.lower() == 'id',
        })
    return profiles


def _infer_entity(table_name, columns):
    """Guess what business entity a table represents."""
    text = (table_name + ' ' + ' '.join(columns)).lower()
    scores = {entity: sum(1 for kw in keywords if kw in text) for entity, keywords in ENTITY_FINGERPRINTS.items()}
    best = max(scores, key=scores.get)
    return best if scores[best] > 0 else 'unknown'


def _suggest_kpis(tables_info):
    """Generate KPI suggestions based on discovered entities."""
    kpis = []
    entity_types = [t['entity_type'] for t in tables_info]
    if 'customer' in entity_types:
        kpis += ['Customer Retention Rate', 'Customer Lifetime Value', 'Churn Rate']
    if 'order' in entity_types or 'financial' in entity_types:
        kpis += ['Revenue Growth', 'Average Order Value', 'Profit Margin', 'Budget Utilization']
    if 'employee' in entity_types:
        kpis += ['Headcount', 'Employee Retention', 'Avg Salary', 'Attrition Risk']
    if 'product' in entity_types:
        kpis += ['Inventory Turnover', 'Stock Accuracy', 'Demand Forecast']
    if 'project' in entity_types:
        kpis += ['Task Completion Rate', 'Project On-Time Delivery', 'Budget vs Actual']
    return list(set(kpis))


def _build_report(tables_info, relationships, filename, record_counts):
    dept_set = set()
    for t in tables_info:
        dept = DEPT_SUGGESTIONS.get(t['entity_type'])
        if dept:
            dept_set.add(dept)

    suggested_depts = list(dept_set) or ['Operations']
    kpi_suggestions = _suggest_kpis(tables_info)

    total_records = sum(record_counts.values())
    insights = []
    if total_records > 10000:
        insights.append(f'📊 Large dataset detected ({total_records:,} records) — consider indexing key columns.')
    if len(relationships) > 3:
        insights.append(f'🔗 {len(relationships)} entity relationships found — strong relational structure.')
    high_null = [t['table'] for t in tables_info if any(c['null_pct'] > 30 for c in t['columns'])]
    if high_null:
        insights.append(f'⚠️ High null rate in: {", ".join(high_null[:3])} — data quality review recommended.')

    return {
        'filename': filename,
        'tables': tables_info,
        'relationships': relationships,
        'record_counts': record_counts,
        'total_records': total_records,
        'suggested_departments': suggested_depts,
        'suggested_kpis': kpi_suggestions,
        'migration_effort': 'High' if len(tables_info) > 10 else ('Medium' if len(tables_info) > 5 else 'Low'),
        'data_quality_score': round(100 - _mean([c['null_pct'] for t in tables_info for c in t['columns']]), 1),
        'insights': insights,
        'recommendation': f'This database maps to {len(suggested_depts)} departments: {", ".join(suggested_depts)}. '
                          f'Estimated migration effort: {"High" if len(tables_info) > 10 else "Low-Medium"}.'
    }


def _analyze_sqlite(file_bytes):
    """Analyze a SQLite .db file."""
    import tempfile, os
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        conn = sqlite3.connect(tmp_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'")
        table_names = [r[0] for r in cur.fetchall()]

        tables_info = []
        record_counts = {}
        for tbl in table_names:
            cur.execute(f"SELECT COUNT(*) FROM [{tbl}]")
            count = cur.fetchone()[0]
            record_counts[tbl] = count
            cur.execute(f"SELECT * FROM [{tbl}] LIMIT 50")
            rows = [dict(r) for r in cur.fetchall()]
            cols = list(rows[0].keys()) if rows else []
            cols_info = _profile_columns(cols, rows)
            entity = _infer_entity(tbl, cols)
            tables_info.append({'table': tbl, 'record_count': count, 'columns': cols_info, 'entity_type': entity, 'column_names': cols})

        # Detect relationships via FK naming
        relationships = []
        for t in tables_info:
            for c in t['columns']:
                if c['is_likely_fk'] and c['column'].lower() != 'id':
                    ref_name = c['column'].replace('_id', '')
                    matching = [other['table'] for other in tables_info if other['table'].lower().startswith(ref_name)]
                    if matching:
                        relationships.append({'from': t['table'], 'column': c['column'], 'references': matching[0]})
        conn.close()
        return _build_report(tables_info, relationships, 'legacy.db', record_counts)
    finally:
        os.unlink(tmp_path)


def _analyze_csv(file_bytes):
    """Analyze a CSV file."""
    text = file_bytes.decode('utf-8', errors='replace')
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)[:500]
    if not rows:
        return {'error': 'CSV file is empty or invalid.'}
    cols = list(rows[0].keys())
    cols_info = _profile_columns(cols, rows)
    entity = _infer_entity('data', cols)
    tables_info = [{'table': 'data', 'record_count': len(rows), 'columns': cols_info, 'entity_type': entity, 'column_names': cols}]
    return _build_report(tables_info, [], 'legacy.csv', {'data': len(rows)})


def _analyze_excel(file_bytes):
    """Analyze an Excel (.xlsx) file."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        tables_info = []
        record_counts = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue
            headers = [str(h) if h is not None else f'col_{i}' for i, h in enumerate(all_rows[0])]
            data_rows = [dict(zip(headers, row)) for row in all_rows[1:501]]
            record_counts[sheet_name] = len(all_rows) - 1
            cols_info = _profile_columns(headers, data_rows)
            entity = _infer_entity(sheet_name, headers)
            tables_info.append({'table': sheet_name, 'record_count': len(data_rows), 'columns': cols_info, 'entity_type': entity, 'column_names': headers})
        wb.close()
        return _build_report(tables_info, [], 'legacy.xlsx', record_counts)
    except ImportError:
        return {'error': 'openpyxl not installed. Run: pip install openpyxl'}
    except Exception as e:
        return {'error': f'Failed to parse Excel file: {str(e)}'}
